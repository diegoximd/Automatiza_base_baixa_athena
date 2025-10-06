import pandas as pd
from openpyxl import Workbook
import firebirdsql
import os
from datetime import datetime

# === Configurações ===
arquivo_excel_origem = r"C:\Users\Suporte\Desktop\Mercado\Athena Saude\PAGAMENTO HUMANA SUL - 01A15092025.xlsx"
arquivo_excel_saida = r"C:\Users\Suporte\Desktop\Mercado\Athena Saude\HUMANA_SUL_ARQUIVO_BAIXA_800_ATHENA.xlsx"
arquivo_sql_saida = r"C:\Users\Suporte\Desktop\Mercado\Athena Saude\HUMANA_SUL_ARQUIVO_BAIXA_800_ATHENAs.sql"
banco_id = 2004

# === Conexão com o Firebird ===
conn = firebirdsql.connect(
    host='',
    database=r'',
    port=3050,
    user='',
    password=''
)

# Consulta operações válidas no banco (ignora STATUS='L') sem gerar warning
cursor = conn.cursor()
cursor.execute(f"SELECT NROPERACAO FROM OPERACOES WHERE BANCO = {banco_id} AND STATUS <> 'L'")
dados = cursor.fetchall()
df_banco = pd.DataFrame(dados, columns=["NROPERACAO"])
df_banco["NROPERACAO"] = df_banco["NROPERACAO"].astype(str)
cursor.close()

# === Leitura do Excel de origem ===
df_origem = pd.read_excel(arquivo_excel_origem, dtype=str)
df_origem.columns = df_origem.columns.str.strip()

# Filtra apenas as operações que existem no banco
df_origem = df_origem[df_origem["Documento"].astype(str).isin(df_banco["NROPERACAO"])]

# === Mapeamento ===
mapeamento = {
    "Documento": "NR OPERAÇÃO",
    "Mensalidade (R$)": "VALOR VENCIDO",
    "Vencimento": "DT. VENCIMENTO",
    "Pagamento": "DT. PAGAMENTO",
    "Valor pago (R$)": "VALOR PAGO",
    "CPF do titular": "CPF / CNPJ",
    "Titular": "NOME DO CLIENTE",
}

cols = [c for c in mapeamento if c in df_origem.columns]
df_convertido = df_origem[cols].rename(columns={k: mapeamento[k] for k in cols})

# Formatação de datas e valores
for c in ["DT. VENCIMENTO", "DT. PAGAMENTO"]:
    if c in df_convertido:
        df_convertido[c] = pd.to_datetime(df_convertido[c], errors="coerce").dt.strftime("%d/%m/%Y")

for c in ["VALOR VENCIDO", "VALOR PAGO"]:
    if c in df_convertido:
        df_convertido[c] = pd.to_numeric(df_convertido[c], errors="coerce").map(
            lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(x) else ""
        )

df_convertido["TIPO"] = "1"
df_convertido["CPF / CNPJ"] = df_convertido.get("CPF / CNPJ", "").map(
    lambda x: str(x).strip() if pd.notna(x) and len(str(x).strip()) > 11 else ""
)

# === Cria Excel de baixa ===
wb = Workbook()
ws = wb.active
ws.title = "BAIXA"

ws.cell(row=1, column=1, value="Dt. Remessa")
ws.cell(row=1, column=2, value="Número da Remessa")
ws.cell(row=1, column=3, value="Código da Empresa")

# Datas e remessa
nome_arquivo = os.path.basename(arquivo_excel_origem)
data_str = ''.join(filter(str.isdigit, nome_arquivo[-12:-4]))
data_formatada = datetime.strptime(data_str, "%d%m%Y").strftime("%d/%m/%Y")
remessa = f"{banco_id}{datetime.strptime(data_str, '%d%m%Y').strftime('%y%m%d')}"
ws.cell(row=2, column=1, value=data_formatada)
ws.cell(row=2, column=2, value=remessa)
ws.cell(row=2, column=3, value=banco_id)

header_detalhe = [
    "TIPO", "NR OPERAÇÃO", "NOME OPERAÇÃO", "VALOR VENCIDO", "DT. VENCIMENTO",
    "DT. PAGAMENTO", "VALOR PAGO", "VALOR COMISSAO", "CPF / CNPJ", "NOME DO CLIENTE",
    "VALOR NOMINAL", "VALOR JUROS", "VALOR DESCONTO", "VALOR MULTA"
]
for col_idx, nome in enumerate(header_detalhe, start=1):
    ws.cell(row=3, column=col_idx, value=nome)

for r_idx, (_, serie) in enumerate(df_convertido.iterrows(), start=4):
    for col_idx, nome in enumerate(header_detalhe, start=1):
        val = serie.get(nome, "")
        ws.cell(row=r_idx, column=col_idx, value=val)

wb.save(arquivo_excel_saida)

# === Gera arquivo SQL ===
with open(arquivo_sql_saida, "w", encoding="utf-8") as f:
    for _, row in df_origem.iterrows():
        nro = row["Documento"]
        venc = row["Vencimento"]
        if pd.notna(nro) and pd.notna(venc):
            f.write(f"UPDATE OPERACOES SET STATUS = 'L' WHERE NROPERACAO = '{nro}' AND DATAVENCTO = '{venc}' AND BANCO = {banco_id};\n")

print("✅ Excel e SQL gerados apenas com operações existentes no banco.")
