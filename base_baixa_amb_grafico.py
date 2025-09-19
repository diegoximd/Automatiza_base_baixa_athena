import pandas as pd
import datetime
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import firebirdsql
from openpyxl import Workbook
import os

# Função para dividir telefones
def split_phones(phone_str):
    if pd.isna(phone_str) or not isinstance(phone_str, str) or not phone_str.strip():
        return ['', '', '', '', '', '']
    phones = [p.strip() for p in phone_str.split('|') if p.strip()]
    phones += [''] * (6 - len(phones))
    return phones[:6]

# Função para criar o DataFrame para o arquivo de base
def create_base_df(input_df):
    model_columns = [
        'TIPO', 'NR OPERAÇÃO', 'NOME OPERAÇÃO', 'AGENCIA', 'CONTA', 'PRODUTO', 'MODALIDADE',
        'DT. ATUALIZADO', 'DT. VENCIMENTO', 'VALOR OPERAÇÃO', 'VALOR VENCIDO', 'VALOR IOF',
        'COND. NEGOCIAIS', 'FORMA ATUALIZAÇÃO', 'GARANTIAS', 'NR. IDENTIDADE', 'CPF / CNPJ',
        'MCI', 'NR FICHA', 'NOME DO CLIENTE', 'ENDEREÇO', 'NUMERO', 'BAIRRO', 'CEP', 'CIDADE',
        'UF', 'TELEFONE 1', 'TELEFONE 2', 'TELEFONE 3', 'TELEFONE 4', 'TELEFONE 5', 'TELEFONE 6',
        'DATA NASCIMENTO', 'NATURALIDADE', 'SEXO', 'ESTADO CIVIL', 'NOME DO PAI', 'NOME DA MÃE',
        'NOME AVALISTA 1', 'CPF/CNPJ AVALISTA 1', 'ENDEREÇO AVALISTA 1', 'BAIRRO AVALISTA 1',
        'CEP AVALISTA 1', 'CIDADE AVALISTA 1', 'UF AVALISTA 1', 'TELEFONE 1 AVALISTA 1',
        'TELEFONE 2 AVALISTA 1', 'NOME AVALISTA 2', 'CPF/CNPJ AVALISTA 2', 'ENDEREÇO AVALISTA 2',
        'BAIRRO AVALISTA 2', 'CEP AVALISTA 2', 'CIDADE AVALISTA 2', 'UF AVALISTA 2',
        'TELEFONE 1 AVALISTA 2', 'TELEFONE 2 AVALISTA 2', 'NOME AVALISTA 3', 'CPF/CNPJ AVALISTA 3',
        'ENDEREÇO AVALISTA 3', 'BAIRRO AVALISTA 3', 'CEP AVALISTA 3', 'CIDADE AVALISTA 3',
        'UF AVALISTA 3', 'TELEFONE 1 AVALISTA 3', 'TELEFONE 2 AVALISTA 3', 'NOME AVALISTA 4',
        'CPF/CNPJ AVALISTA 4', 'ENDEREÇO AVALISTA 4', 'BAIRRO AVALISTA 4', 'CEP AVALISTA 4',
        'CIDADE AVALISTA 4', 'UF AVALISTA 4', 'TELEFONE 1 AVALISTA 4', 'TELEFONE 2 AVALISTA 4',
        'NOME AVALISTA 5', 'CPF/CNPJ AVALISTA 5', 'ENDEREÇO AVALISTA 5', 'BAIRRO AVALISTA 5',
        'CEP AVALISTA 5', 'CIDADE AVALISTA 5', 'UF AVALISTA 5', 'TELEFONE 1 AVALISTA 5',
        'TELEFONE 2 AVALISTA 5', 'NOME AVALISTA 6', 'CPF/CNPJ AVALISTA 6', 'ENDEREÇO AVALISTA 6',
        'BAIRRO AVALISTA 6', 'CEP AVALISTA 6', 'CIDADE AVALISTA 6', 'UF AVALISTA 6',
        'TELEFONE 1 AVALISTA 6', 'TELEFONE 2 AVALISTA 6', 'PROFISSÃO', 'NOME LOCAL DE TRABALHO',
        'ENDEREÇO LOCAL DE TRABALHO', 'BAIRRO LOCAL DE TRABALHO', 'CEP LOCAL DE TRABALHO',
        'CIDADE LOCAL DE TRABALHO', 'UF LOCAL DE TRABALHO', 'TELEFONE 1 LOCAL DE TRABALHO',
        'TELEFONE 2 LOCAL DE TRABALHO', 'REFERENCIA PESSOAL', 'TELEFONE 1 REFERENCIA',
        'TELEFONE 2 REFERENCIA', 'REFERENCIA PESSOAL 2', 'TELEFONE 1 REFERENCIA 2',
        'TELEFONE 2 REFERENCIA 2', 'REFERENCIA PESSOAL 3', 'TELEFONE 1 REFERENCIA 3',
        'TELEFONE 2 REFERENCIA 3', 'SPC/SERASA', 'E-MAIL', 'E-MAIL1', 'DT. EMISSÃO',
        'VALOR PROTESTO', 'OBS. OPERAÇÃO', 'OBS. CLIENTE', 'DT. FIMTERCERIZAÇÃO', 'VALOR JUROS',
        'COD_CLASSIFICACAO_CLIENTE', 'COD_CLASSIFICACAO_OPERACAO', 'DATA ASSINATURA DO CONTRATO',
        'SCORE', 'SCORE INTERNO', 'RENDA'
    ]

    output_df = pd.DataFrame(columns=model_columns)

    if not input_df.empty:
        output_df['NR OPERAÇÃO'] = input_df['DOCUMENTO'].astype(str)
        output_df['NOME OPERAÇÃO'] = input_df['PLANO']
        output_df['AGENCIA'] = input_df['LOCAL_PAGAMENTO']
        output_df['CONTA'] = input_df['CODIGO'].astype(str)
        output_df['PRODUTO'] = input_df['TIPO PLANO']
        output_df['MODALIDADE'] = input_df['MATRIZ DE OFERTA']
        output_df['VALOR OPERAÇÃO'] = input_df['VALOR_TOTAL']
        output_df['CPF / CNPJ'] = input_df['CPF_CNPJ_CAEPF'].astype(str)
        output_df['NOME DO CLIENTE'] = input_df['TITULAR']
        output_df['ENDEREÇO'] = input_df['ENDERECO']
        output_df['NUMERO'] = input_df['NUMERO']
        output_df['BAIRRO'] = input_df['BAIRRO']
        output_df['CEP'] = input_df['CEP']
        output_df['CIDADE'] = input_df['CIDADE']
        output_df['UF'] = input_df['ESTADO']
        output_df['E-MAIL'] = input_df['EMAIL']
        output_df['COND. NEGOCIAIS'] = input_df['SITUACAO']
        output_df['GARANTIAS'] = input_df['NATUREZA_CONTRATO']
        output_df['FORMA ATUALIZAÇÃO'] = input_df['ESTABELECIMENTO']
        output_df['OBS. OPERAÇÃO'] = input_df['STATUS']

        def format_date(x):
            if pd.isna(x):
                return ''
            if isinstance(x, datetime.datetime):
                return x.strftime('%d/%m/%Y')
            try:
                excel_epoch = datetime.date(1899, 12, 30)
                date_val = excel_epoch + datetime.timedelta(days=int(x))
                return date_val.strftime('%d/%m/%Y')
            except:
                return str(x)

        output_df['DT. VENCIMENTO'] = input_df['VENCIMENTO'].apply(format_date)

        if 'TELEFONE' in input_df.columns:
            phones = input_df['TELEFONE'].astype(str).apply(split_phones)
            output_df['TELEFONE 1'] = phones.apply(lambda x: x[0])
            output_df['TELEFONE 2'] = phones.apply(lambda x: x[1])
            output_df['TELEFONE 3'] = phones.apply(lambda x: x[2])
            output_df['TELEFONE 4'] = phones.apply(lambda x: x[3])
            output_df['TELEFONE 5'] = phones.apply(lambda x: x[4])
            output_df['TELEFONE 6'] = phones.apply(lambda x: x[5])
        else:
            output_df['TELEFONE 1'] = ''
            output_df['TELEFONE 2'] = ''
            output_df['TELEFONE 3'] = ''
            output_df['TELEFONE 4'] = ''
            output_df['TELEFONE 5'] = ''
            output_df['TELEFONE 6'] = ''

        output_df['TIPO'] = '1'
        output_df['TIPO'] = output_df['TIPO'].astype(str).fillna('1')

    return output_df

# Função para criar o DataFrame para o arquivo de baixa
def create_baixa_df(df_origem, banco_id):
    try:
        conn = firebirdsql.connect(
            host='servidor2',
            database=r'D:\Dados_interbase\COB_DB_EXECUTIVA_ATHENA_SAUDE.FDB',
            port=3050,
            user='consulta',
            password='@BmpAdm35ConsultaSql#'
        )
        cursor = conn.cursor()
        cursor.execute(f"SELECT NROPERACAO FROM OPERACOES WHERE BANCO = {banco_id} AND STATUS <> 'L'")
        dados = cursor.fetchall()
        df_banco = pd.DataFrame(dados, columns=["NROPERACAO"])
        df_banco["NROPERACAO"] = df_banco["NROPERACAO"].astype(str)
        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao conectar ao banco de dados: {e}")
        return None, None

    df_origem = df_origem[df_origem["Documento"].astype(str).isin(df_banco["NROPERACAO"])]

    mapeamento = {
        "Documento": "NR OPERAÇÃO",
        "Mensalidade (R$)": "VALOR VENCIDO",
        "Vencimento": "DT. VENCIMENTO",
        "Pagamento": "DT. PAGAMENTO",
        "Valor pago (R$)": "VALOR PAGO",
        "CPF do titular": "CPF / CNPJ",
        "Titular": "NOME DO CLIENTE",
    }

    cols = [c for c in mapeamento if c in df_origem.columns]
    df_convertido = df_origem[cols].rename(columns={k: mapeamento[k] for k in cols})

    for c in ["DT. VENCIMENTO", "DT. PAGAMENTO"]:
        if c in df_convertido:
            df_convertido[c] = pd.to_datetime(df_convertido[c], errors="coerce").dt.strftime("%d/%m/%Y")

    for c in ["VALOR VENCIDO", "VALOR PAGO"]:
        if c in df_convertido:
            df_convertido[c] = pd.to_numeric(df_convertido[c], errors="coerce").map(
                lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(x) else ""
            )

    df_convertido["TIPO"] = "1"
    # Filtrar CPFs com 11 caracteres ou menos
    original_count = len(df_convertido)
    df_convertido["CPF / CNPJ"] = df_convertido.get("CPF / CNPJ", "").map(
        lambda x: str(x).strip() if pd.notna(x) and len(str(x).strip()) > 11 else ""
    )
    removed_count = original_count - len(df_convertido[df_convertido["CPF / CNPJ"] != ""])

    return df_convertido, df_origem

# Função para escrever o arquivo de base
def write_base_excel(output_file, model_df, company_code, remessa_num, excel_serial):
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        initial_header = pd.DataFrame(
            [['Dt. Remessa', 'Número da Remessa', 'Código da Empresa', 'Código de Evento Ref. A Atualização',
              'Retomar/Liquidar Operacao não Presentes'] + [''] * (len(model_df.columns) - 5)])
        initial_header.to_excel(writer, sheet_name='Modelo_Excel_Incluir_Clientes_I', index=False, header=False, startrow=0)

        numero_remessa = company_code + remessa_num
        row2_data = [excel_serial, numero_remessa, company_code, '', 'RETOMAR'] + [''] * (len(model_df.columns) - 5)
        pd.DataFrame([row2_data]).to_excel(writer, sheet_name='Modelo_Excel_Incluir_Clientes_I', index=False, header=False, startrow=1)

        pd.DataFrame([model_df.columns]).to_excel(writer, sheet_name='Modelo_Excel_Incluir_Clientes_I', index=False, header=False, startrow=2)

        model_df.to_excel(writer, sheet_name='Modelo_Excel_Incluir_Clientes_I', index=False, header=False, startrow=3)

# Função para escrever o arquivo de baixa
def write_baixa_excel(output_file, df_convertido, banco_id, data_formatada, remessa):
    wb = Workbook()
    ws = wb.active
    ws.title = "BAIXA"

    ws.cell(row=1, column=1, value="Dt. Remessa")
    ws.cell(row=1, column=2, value="Número da Remessa")
    ws.cell(row=1, column=3, value="Código da Empresa")

    ws.cell(row=2, column=1, value=data_formatada)
    ws.cell(row=2, column=2, value=remessa)
    ws.cell(row=2, column=3, value=banco_id)

    header_detalhe = [
        "TIPO", "NR OPERAÇÃO", "NOME OPERAÇÃO", "VALOR VENCIDO", "DT. VENCIMENTO",
        "DT. PAGAMENTO", "VALOR PAGO", "VALOR COMISSAO", "CPF / CNPJ", "NOME DO CLIENTE",
        "VALOR NOMINAL", "VALOR JUROS", "VALOR DESCONTO", "VALOR MULTA"
    ]
    for col_idx, nome in enumerate(header_detalhe, start=1):
        ws.cell(row=3, column=col_idx, value=nome)

    for r_idx, (_, serie) in enumerate(df_convertido.iterrows(), start=4):
        for col_idx, nome in enumerate(header_detalhe, start=1):
            val = serie.get(nome, "")
            ws.cell(row=r_idx, column=col_idx, value=val)

    wb.save(output_file)

# Função para gerar o arquivo SQL
def generate_sql_file(output_sql, df_origem, banco_id):
    with open(output_sql, "w", encoding="utf-8") as f:
        for _, row in df_origem.iterrows():
            nro = row["Documento"]
            venc = row["Vencimento"]
            if pd.notna(nro) and pd.notna(venc):
                try:
                    # Formatar a data no padrão americano aaaa-mm-dd
                    venc = pd.to_datetime(venc).strftime('%Y-%m-%d')
                    f.write(
                        f"UPDATE OPERACOES SET STATUS = 'L' "
                        f"WHERE NROPERACAO = '{nro}' AND DATAVENCTO = '{venc}' AND BANCO = {banco_id};\n"
                    )
                except Exception as e:
                    # Logar erro, mas continuar processando outros registros
                    print(f"Erro ao formatar data {venc} para NROPERACAO {nro}: {e}")

# Interface gráfica com Layout 2
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de Arquivos Base e Baixa")
        self.root.geometry("700x450")
        self.root.configure(bg="#f0f0f0")  # Fundo cinza claro

        # Variáveis
        self.source_file = tk.StringVar()
        self.output_type = tk.StringVar(value="Base")
        self.company = tk.StringVar(value="Humana NE")
        self.output_dir = tk.StringVar(value=r"C:\Users\Suporte\Desktop\Mercado\Athena Saude")

        # Painel de título
        title_frame = tk.Frame(root, bg="#003087")  # Azul escuro
        title_frame.pack(fill="x", pady=10)
        tk.Label(title_frame, text="Gerador de Arquivos Base e Baixa", font=("Arial", 16, "bold"),
                 bg="#003087", fg="white").pack(pady=10)

        # Painel de seleção de arquivo
        file_frame = tk.Frame(root, bg="#f0f0f0", bd=1, relief="groove")
        file_frame.pack(padx=15, pady=10, fill="x")
        tk.Label(file_frame, text="Arquivo de Origem:", font=("Arial", 12, "bold"), bg="#f0f0f0").pack(anchor="w", padx=10)
        tk.Entry(file_frame, textvariable=self.source_file, width=50).pack(side="left", padx=10, pady=5)
        tk.Button(file_frame, text="Procurar", command=self.browse_file, bg="#0052cc", fg="white").pack(side="left", padx=10)

        # Painel de opções (Tipo de Arquivo e Empresa)
        options_frame = tk.Frame(root, bg="#f0f0f0", bd=1, relief="groove")
        options_frame.pack(padx=15, pady=10, fill="x")

        # Subpainel para Tipo de Arquivo
        type_frame = tk.Frame(options_frame, bg="#f0f0f0")
        type_frame.pack(side="left", padx=10, fill="y")
        tk.Label(type_frame, text="Tipo de Arquivo:", font=("Arial", 12, "bold"), bg="#f0f0f0").pack(anchor="w")
        tk.Radiobutton(type_frame, text="Base", variable=self.output_type, value="Base",
                       command=self.toggle_company_selection, bg="#f0f0f0").pack(anchor="w", pady=5)
        tk.Radiobutton(type_frame, text="Baixa", variable=self.output_type, value="Baixa",
                       command=self.toggle_company_selection, bg="#f0f0f0").pack(anchor="w")

        # Subpainel para Empresa
        self.company_frame = tk.Frame(options_frame, bg="#f0f0f0")
        self.company_frame.pack(side="left", padx=10, fill="y")
        tk.Label(self.company_frame, text="Empresa:", font=("Arial", 12, "bold"), bg="#f0f0f0").pack(anchor="w")
        self.company_combo = ttk.Combobox(self.company_frame, textvariable=self.company,
                                         values=["Humana NE", "SAMP", "Humana SUL"], state="disabled")
        self.company_combo.pack(pady=5)

        # Painel de seleção de diretório de destino
        dir_frame = tk.Frame(root, bg="#f0f0f0", bd=1, relief="groove")
        dir_frame.pack(padx=15, pady=10, fill="x")
        tk.Label(dir_frame, text="Diretório de Destino:", font=("Arial", 12, "bold"), bg="#f0f0f0").pack(anchor="w", padx=10)
        tk.Entry(dir_frame, textvariable=self.output_dir, width=50).pack(side="left", padx=10, pady=5)
        tk.Button(dir_frame, text="Selecionar Destino", command=self.browse_directory, bg="#0052cc", fg="white").pack(side="left", padx=10)

        # Painel do botão de ação
        action_frame = tk.Frame(root, bg="#f0f0f0")
        action_frame.pack(pady=20)
        tk.Button(action_frame, text="Gerar Arquivo(s)", command=self.generate_files,
                  font=("Arial", 12), bg="#006400", fg="white").pack()

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.source_file.set(file_path)

    def browse_directory(self):
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.output_dir.set(dir_path)

    def toggle_company_selection(self):
        if self.output_type.get() == "Baixa":
            self.company_combo.config(state="normal")
        else:
            self.company_combo.config(state="disabled")

    def generate_files(self):
        source_file = self.source_file.get()
        output_type = self.output_type.get()
        company = self.company.get()
        output_dir = self.output_dir.get()

        if not source_file:
            messagebox.showerror("Erro", "Selecione um arquivo de origem!")
            return

        # Verificar se o diretório de destino existe
        if not os.path.exists(output_dir):
            messagebox.showerror("Erro", f"O diretório de destino '{output_dir}' não existe! Selecione um diretório válido.")
            dir_path = filedialog.askdirectory()
            if not dir_path:
                return
            self.output_dir.set(dir_path)
            output_dir = dir_path

        try:
            if output_type == "Base":
                df = pd.read_excel(source_file, sheet_name='Planilha1', header=0, dtype={'CPF_CNPJ_CAEPF': str})
                df.columns = df.columns.str.strip()

                date_match = re.search(r'(\d{8})\.xlsx$', source_file)
                if date_match:
                    date_str = date_match.group(1)
                    remessa_date = datetime.datetime.strptime(date_str, '%d%m%Y').date()
                else:
                    messagebox.showerror("Erro", "Não foi possível extrair a data do nome do arquivo.")
                    return

                remessa_num = remessa_date.strftime('%y%m%d')
                excel_serial = remessa_date.strftime('%d/%m/%Y')

                df['ESTABELECIMENTO_norm'] = df['ESTABELECIMENTO'].str.replace(r'\s+', ' ', regex=True).str.strip()

                ne_keywords = [
                    'HUMANA ASSISTENCIA MEDICA - THE', 'HUMANA ASSISTENCIA MEDICA - MA',
                    'HUMANA ASSISTENCIA MEDICA - PHB', 'HUMANA ASSISTENCIA MEDICA - FLORIANO',
                    'HUMANA ASSISTENCIA MEDICA - PICOS', 'MEDPLAN', 'HUMANA ASSISTENCIA MEDICA - UNIHOSP',
                    'HUMANA ASSISTENCIA MEDICA - NATAL', 'HUMANA ASSISTENCIA MEDICA - ONCOLIFE',
                    'HUMANA SAUDE NORDESTE - CLINICA ABA TERESINA II FREI SERAFIM'
                ]
                ne_keywords_norm = [re.sub(r'\s+', ' ', kw).strip() for kw in ne_keywords]

                samp_keywords = [
                    '007-00 SAMP ESPIRITO SANTO ASSISTENCIA MEDICA LTDA',
                    '099-SAO BERNARDO SAUDE', '114 - CLINICA ABA SERRA'
                ]
                samp_keywords_norm = [re.sub(r'\s+', ' ', kw).strip() for kw in samp_keywords]

                sul_keywords = [
                    'HUMANA SAUDE SUL LTDA – MARINGA E REGIAO', 'HUMANA SAUDE SUL LTDA – CAXIAS',
                    'HUMANA SAUDE SUL LTDA – RONDON', 'HUMANA SAUDE SUL LTDA – HOSPITAL RONDON',
                    'HUMANA SAUDE SUL LTDA – CAXIAS – MEDICINA OCUPACIONAL'
                ]
                sul_keywords_norm = [re.sub(r'\s+', ' ', kw).strip() for kw in sul_keywords]

                df_ne = df[df['ESTABELECIMENTO_norm'].str.contains('|'.join(ne_keywords_norm), case=False, na=False)]
                df_samp = df[df['ESTABELECIMENTO_norm'].str.contains('|'.join(samp_keywords_norm), case=False, na=False)]
                df_sul = df[df['ESTABELECIMENTO_norm'].str.contains('|'.join(sul_keywords_norm), case=False, na=False)]

                df_ne_model = create_base_df(df_ne)
                df_samp_model = create_base_df(df_samp)
                df_sul_model = create_base_df(df_sul)

                write_base_excel(
                    os.path.join(output_dir, 'HUMANA_NE_ARQUIVO_BASE_800_ATHENA.xlsx'),
                    df_ne_model, '2002', remessa_num, excel_serial
                )
                write_base_excel(
                    os.path.join(output_dir, 'SAMP_ARQUIVO_BASE_800_ATHENA.xlsx'),
                    df_samp_model, '2003', remessa_num, excel_serial
                )
                write_base_excel(
                    os.path.join(output_dir, 'HUMANA_SUL_ARQUIVO_BASE_800_ATHENA.xlsx'),
                    df_sul_model, '2004', remessa_num, excel_serial
                )

                messagebox.showinfo("Sucesso", "Arquivos de base gerados com sucesso!")

            else:  # Baixa
                company_map = {
                    "Humana NE": (2002, "HUMANA_NE_ARQUIVO_BAIXA_800_ATHENA.xlsx", "HUMANA_NE_ARQUIVO_BAIXA_800_ATHENA.sql"),
                    "SAMP": (2003, "SAMP_ARQUIVO_BAIXA_800_ATHENA.xlsx", "SAMP_ARQUIVO_BAIXA_800_ATHENA.sql"),
                    "Humana SUL": (2004, "HUMANA_SUL_ARQUIVO_BAIXA_800_ATHENA.xlsx", "HUMANA_SUL_ARQUIVO_BAIXA_800_ATHENA.sql")
                }
                banco_id, output_excel, output_sql = company_map[company]

                df_origem = pd.read_excel(source_file, dtype=str)
                df_origem.columns = df_origem.columns.str.strip()

                df_convertido, df_origem = create_baixa_df(df_origem, banco_id)
                if df_convertido is None:
                    return

                nome_arquivo = os.path.basename(source_file)
                match = re.search(r"(\d{8})", nome_arquivo)
                if not match:
                    messagebox.showerror("Erro", "Não foi encontrada uma data (ddmmaaaa) no nome do arquivo!")
                    return

                data_str = match.group(1)
                data_formatada = datetime.datetime.strptime(data_str, "%d%m%Y").strftime("%d/%m/%Y")
                remessa = f"{banco_id}{datetime.datetime.strptime(data_str, '%d%m%Y').strftime('%y%m%d')}"

                write_baixa_excel(
                    os.path.join(output_dir, output_excel),
                    df_convertido, banco_id, data_formatada, remessa
                )
                generate_sql_file(
                    os.path.join(output_dir, output_sql),
                    df_origem, banco_id
                )

                messagebox.showinfo("Sucesso", f"Arquivo de baixa e SQL gerados para {company}! {len(df_convertido)} operações exportadas.")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Inicializar a interface gráfica
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()